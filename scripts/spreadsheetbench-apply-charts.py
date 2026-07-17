from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import posixpath
import re
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    BubbleChart,
    DoughnutChart,
    LineChart,
    PieChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.cell import get_column_letter, range_boundaries


CONTRACT_VERSION = 2
MAX_OPERATIONS = 32
MAX_SERIES = 12
MAX_POINTS = 100_000
MAX_ROW = 1_048_576
MAX_COLUMN = 16_384
CHART_TYPES = {"line", "bar", "column", "pie", "doughnut", "scatter", "area", "bubble", "combo"}
COMBO_SERIES_TYPES = {"line", "bar", "column", "area"}
LEGEND_POSITIONS = {"top", "bottom", "left", "right", "none"}
GROUPINGS = {"clustered", "stacked", "percentStacked"}
OPERATION_KEYS = {
    "op", "sheet", "chartType", "title", "categoryRange", "series", "anchor",
    "width", "height", "legendPosition", "grouping", "dataLabels",
}
SERIES_KEYS = {
    "name", "valuesRange", "chartType", "xValuesRange", "sizeRange", "color", "secondaryAxis",
}
CHART_ELEMENT_NAMES = {
    "lineChart", "barChart", "pieChart", "doughnutChart", "scatterChart", "areaChart", "bubbleChart",
}
RANGE_PATTERN = re.compile(
    r"^(?:(?:'((?:[^']|'')+)'|([^'!\[\]]+))!)?"
    r"\$?([A-Z]{1,3})\$?([1-9][0-9]*):\$?([A-Z]{1,3})\$?([1-9][0-9]*)$",
    re.IGNORECASE,
)
CELL_PATTERN = re.compile(r"^([A-Z]{1,3})([1-9][0-9]*)$", re.IGNORECASE)
RELATIONSHIP_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


class ChartContractError(ValueError):
    pass


@dataclass(frozen=True)
class SourceRange:
    input: str
    sheet: str
    address: str
    formula: str
    min_col: int
    min_row: int
    max_col: int
    max_row: int
    point_count: int
    non_empty_count: int

    def reference(self, workbook: Any) -> Reference:
        sheet = workbook[self.sheet]
        return Reference(
            sheet,
            min_col=self.min_col,
            min_row=self.min_row,
            max_col=self.max_col,
            max_row=self.max_row,
        )

    def receipt(self) -> dict[str, Any]:
        return {
            "input": self.input,
            "sheet": self.sheet,
            "address": self.address,
            "formula": self.formula,
            "pointCount": self.point_count,
            "nonEmptyCount": self.non_empty_count,
        }


@dataclass(frozen=True)
class ValidatedSeries:
    name: str
    values: SourceRange
    chart_type: str
    x_values: SourceRange | None
    sizes: SourceRange | None
    color: str | None
    secondary_axis: bool

    def receipt(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "chartType": self.chart_type,
            "valuesRange": self.values.receipt(),
            **({"xValuesRange": self.x_values.receipt()} if self.x_values else {}),
            **({"sizeRange": self.sizes.receipt()} if self.sizes else {}),
            **({"color": self.color} if self.color else {}),
            "secondaryAxis": self.secondary_axis,
        }


@dataclass(frozen=True)
class ValidatedOperation:
    index: int
    chart_type: str
    effective_chart_type: str
    sheet: str
    title: str
    category: SourceRange
    series: tuple[ValidatedSeries, ...]
    anchor: str
    width: float
    height: float
    legend_position: str
    grouping: str
    data_labels: bool

    def receipt(self) -> dict[str, Any]:
        return {
            "index": self.index,
            "chartType": self.chart_type,
            "effectiveChartType": self.effective_chart_type,
            "sheet": self.sheet,
            "title": self.title,
            "categoryRange": self.category.receipt(),
            "series": [item.receipt() for item in self.series],
            "anchor": self.anchor,
            "width": self.width,
            "height": self.height,
            "legendPosition": self.legend_position,
            "grouping": self.grouping,
            "dataLabels": self.data_labels,
        }


def main() -> int:
    parser = argparse.ArgumentParser(description="Apply and verify bounded chart operations on an existing XLSX workbook.")
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--operations", required=True)
    parser.add_argument("--receipt")
    parser.add_argument("--engine", choices=["auto", "excel", "openpyxl"], default="auto")
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    operations_path = Path(args.operations).resolve()
    receipt_path = Path(args.receipt).resolve() if args.receipt else operations_path.with_name("chart-application-receipt.json")
    engine = args.engine
    if engine == "auto":
        engine = "excel" if os.name == "nt" and excel_com_available() else "openpyxl"

    before = empty_package_summary()
    operation_count = 0
    temporary_path: Path | None = None
    validation_workbook: Any = None
    try:
        validate_paths(workbook_path, operations_path, receipt_path)
        payload = read_operation_payload(operations_path)
        raw_operations = payload["operations"]
        chart_operations = [item for item in raw_operations if isinstance(item, dict) and item.get("op") == "add_chart"]
        operation_count = len(chart_operations)
        if not chart_operations:
            raise ChartContractError("operations payload contains no add_chart operations")

        before = inspect_package(workbook_path)
        keep_vba = workbook_path.suffix.lower() == ".xlsm"
        validation_workbook = load_workbook(workbook_path, keep_vba=keep_vba, keep_links=True)
        operations = validate_operations(validation_workbook, chart_operations, before)
        temporary_path = create_temporary_workbook_path(workbook_path)

        if engine == "openpyxl":
            apply_with_openpyxl(validation_workbook, temporary_path, operations)
            validation_workbook = None
        else:
            validation_workbook.close()
            validation_workbook = None
            apply_with_excel_com(workbook_path, temporary_path, operations)

        after = inspect_package(temporary_path)
        operation_receipts = verify_persisted_charts(before, after, operations)
        os.replace(temporary_path, workbook_path)
        temporary_path = None

        receipt = success_receipt(workbook_path, engine, before, after, operations, operation_receipts)
        write_json_atomic(receipt_path, receipt)
        print(json.dumps(receipt, separators=(",", ":"), ensure_ascii=True))
        return 0
    except Exception as error:
        rejected = rejection_receipt(workbook_path, engine, before, operation_count, error)
        try:
            write_json_atomic(receipt_path, rejected)
        except Exception:
            pass
        print(json.dumps(rejected, separators=(",", ":"), ensure_ascii=True), file=sys.stderr)
        return 2
    finally:
        if validation_workbook is not None:
            validation_workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def validate_paths(workbook_path: Path, operations_path: Path, receipt_path: Path) -> None:
    if not workbook_path.is_file():
        raise ChartContractError(f"workbook does not exist: {workbook_path.name}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ChartContractError("chart bridge requires an .xlsx or .xlsm workbook")
    if not operations_path.is_file():
        raise ChartContractError(f"operations payload does not exist: {operations_path.name}")
    if receipt_path in {workbook_path, operations_path}:
        raise ChartContractError("receipt path must differ from workbook and operations paths")


def read_operation_payload(path: Path) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as error:
        raise ChartContractError(f"invalid operations JSON: {error}") from error
    if not isinstance(payload, dict):
        raise ChartContractError("operations payload must be a JSON object")
    if payload.get("schema") not in {1, 2}:
        raise ChartContractError("operations payload schema must be 1 or 2")
    operations = payload.get("operations")
    if not isinstance(operations, list):
        raise ChartContractError("operations payload must contain an operations array")
    if len(operations) > MAX_OPERATIONS:
        raise ChartContractError(f"operations payload exceeds the {MAX_OPERATIONS}-operation bound")
    return payload


def excel_com_available() -> bool:
    try:
        import win32com.client  # type: ignore  # noqa: F401
        return True
    except ImportError:
        return False


def validate_operations(
    workbook: Any,
    raw_operations: list[dict[str, Any]],
    before: dict[str, Any],
) -> list[ValidatedOperation]:
    if not 1 <= len(raw_operations) <= MAX_OPERATIONS:
        raise ChartContractError(f"expected 1-{MAX_OPERATIONS} chart operations")
    occupied_anchors = {
        (str(item.get("sheet", "")).casefold(), str(item.get("anchor", "")).upper())
        for item in before["chartObjects"]
    }
    requested_anchors: set[tuple[str, str]] = set()
    validated: list[ValidatedOperation] = []

    for index, operation in enumerate(raw_operations):
        path = f"operations[{index}]"
        reject_unknown_fields(operation, OPERATION_KEYS, path)
        chart_type = require_choice(operation.get("chartType"), CHART_TYPES, f"{path}.chartType")
        sheet_name = require_text(operation.get("sheet"), f"{path}.sheet", 31)
        if sheet_name not in workbook.sheetnames:
            raise ChartContractError(f"{path}.sheet references missing sheet: {sheet_name}")
        target_sheet = workbook[sheet_name]
        if target_sheet.sheet_state != "visible":
            raise ChartContractError(f"{path}.sheet must be visible: {sheet_name}")

        title = require_text(operation.get("title"), f"{path}.title", 160)
        anchor = require_anchor(operation.get("anchor"), f"{path}.anchor")
        anchor_key = (sheet_name.casefold(), anchor)
        if anchor_key in occupied_anchors:
            raise ChartContractError(f"{path}.anchor overlaps an existing chart: {sheet_name}!{anchor}")
        if anchor_key in requested_anchors:
            raise ChartContractError(f"{path}.anchor duplicates another requested chart: {sheet_name}!{anchor}")
        requested_anchors.add(anchor_key)

        legend_position = require_choice(operation.get("legendPosition"), LEGEND_POSITIONS, f"{path}.legendPosition")
        grouping = require_choice(operation.get("grouping", "clustered"), GROUPINGS, f"{path}.grouping")
        width = require_dimension(operation.get("width"), 18, 6, 36, f"{path}.width")
        height = require_dimension(operation.get("height"), 10, 4, 24, f"{path}.height")
        data_labels = operation.get("dataLabels", False)
        if not isinstance(data_labels, bool):
            raise ChartContractError(f"{path}.dataLabels must be boolean")

        category = parse_source_range(workbook, operation.get("categoryRange"), sheet_name, f"{path}.categoryRange")
        raw_series = operation.get("series")
        if not isinstance(raw_series, list) or not 1 <= len(raw_series) <= MAX_SERIES:
            raise ChartContractError(f"{path}.series must contain 1-{MAX_SERIES} entries")

        has_override = any(
            isinstance(item, dict) and (item.get("chartType") is not None or item.get("secondaryAxis") is True)
            for item in raw_series
        )
        effective_chart_type = "combo" if chart_type == "combo" or has_override else chart_type
        if effective_chart_type == "combo" and len(raw_series) < 2:
            raise ChartContractError(f"{path}.series: combo charts require at least two series")
        if chart_type in {"pie", "doughnut"} and len(raw_series) != 1:
            raise ChartContractError(f"{path}.series: {chart_type} charts require exactly one series")

        series: list[ValidatedSeries] = []
        for series_index, item in enumerate(raw_series):
            series_path = f"{path}.series[{series_index}]"
            if not isinstance(item, dict):
                raise ChartContractError(f"{series_path} must be an object")
            reject_unknown_fields(item, SERIES_KEYS, series_path)
            name = require_text(item.get("name"), f"{series_path}.name", 160)
            values = parse_source_range(workbook, item.get("valuesRange"), sheet_name, f"{series_path}.valuesRange")

            declared_type = item.get("chartType")
            if declared_type is not None:
                declared_type = require_choice(declared_type, COMBO_SERIES_TYPES, f"{series_path}.chartType")
            if chart_type == "combo" and declared_type is None:
                raise ChartContractError(f"{series_path}.chartType is required for explicit combo charts")
            if effective_chart_type == "combo":
                fallback_type = chart_type if chart_type in COMBO_SERIES_TYPES else None
                actual_type = declared_type or fallback_type
                if actual_type is None:
                    raise ChartContractError(f"{series_path}.chartType is required for combo charts")
            else:
                if declared_type is not None or item.get("secondaryAxis") is not None:
                    raise ChartContractError(f"{series_path}: combo metadata requires a combo-compatible chart")
                actual_type = chart_type

            secondary_axis = item.get("secondaryAxis", False)
            if not isinstance(secondary_axis, bool):
                raise ChartContractError(f"{series_path}.secondaryAxis must be boolean")
            color = validate_color(item.get("color"), f"{series_path}.color")

            x_values: SourceRange | None = None
            sizes: SourceRange | None = None
            if chart_type in {"scatter", "bubble"}:
                if item.get("xValuesRange") is None:
                    raise ChartContractError(f"{series_path}.xValuesRange is required for {chart_type} charts")
                x_values = parse_source_range(workbook, item.get("xValuesRange"), sheet_name, f"{series_path}.xValuesRange")
                require_same_point_count(x_values, values, f"{series_path}.valuesRange")
            elif item.get("xValuesRange") is not None:
                raise ChartContractError(f"{series_path}.xValuesRange is only valid for scatter and bubble charts")
            else:
                require_same_point_count(category, values, f"{series_path}.valuesRange")

            if chart_type == "bubble":
                if item.get("sizeRange") is None:
                    raise ChartContractError(f"{series_path}.sizeRange is required for bubble charts")
                sizes = parse_source_range(workbook, item.get("sizeRange"), sheet_name, f"{series_path}.sizeRange")
                require_same_point_count(x_values, sizes, f"{series_path}.sizeRange")
                validate_bubble_sizes(workbook, sizes, f"{series_path}.sizeRange")
            elif item.get("sizeRange") is not None:
                raise ChartContractError(f"{series_path}.sizeRange is only valid for bubble charts")

            series.append(ValidatedSeries(
                name=name,
                values=values,
                chart_type=actual_type,
                x_values=x_values,
                sizes=sizes,
                color=color,
                secondary_axis=secondary_axis,
            ))

        if effective_chart_type == "combo":
            if not any(not item.secondary_axis for item in series):
                raise ChartContractError(f"{path}.series: combo charts require a primary-axis series")
            distinct_types = {item.chart_type for item in series}
            if len(distinct_types) < 2 and not any(item.secondary_axis for item in series):
                raise ChartContractError(f"{path}.series: combo charts need multiple chart types or a secondary axis")

        validated.append(ValidatedOperation(
            index=index,
            chart_type=chart_type,
            effective_chart_type=effective_chart_type,
            sheet=sheet_name,
            title=title,
            category=category,
            series=tuple(series),
            anchor=anchor,
            width=width,
            height=height,
            legend_position=legend_position,
            grouping=grouping,
            data_labels=data_labels,
        ))
    return validated


def parse_source_range(workbook: Any, value: Any, default_sheet: str, path: str) -> SourceRange:
    if not isinstance(value, str):
        raise ChartContractError(f"{path} must be a one-dimensional A1 range")
    stripped = value.strip().lstrip("=")
    if not stripped or any(token in stripped for token in ("[", "]", ",")):
        raise ChartContractError(f"{path} cannot be external, union, named, or empty")
    match = RANGE_PATTERN.fullmatch(stripped)
    if not match:
        raise ChartContractError(f"{path} must be a contiguous A1 range such as 'Data'!A2:A13")
    sheet_name = (match.group(1).replace("''", "'") if match.group(1) else (match.group(2) or default_sheet).strip())
    if sheet_name not in workbook.sheetnames:
        raise ChartContractError(f"{path} references missing source sheet: {sheet_name}")
    min_col, min_row, max_col, max_row = range_boundaries(
        f"{match.group(3)}{match.group(4)}:{match.group(5)}{match.group(6)}"
    )
    if min_col > max_col or min_row > max_row or max_col > MAX_COLUMN or max_row > MAX_ROW:
        raise ChartContractError(f"{path} is outside Excel worksheet bounds")
    if min_col != max_col and min_row != max_row:
        raise ChartContractError(f"{path} must be one-dimensional, not a two-dimensional block")
    point_count = (max_col - min_col + 1) * (max_row - min_row + 1)
    if point_count > MAX_POINTS:
        raise ChartContractError(f"{path} exceeds the {MAX_POINTS}-point bound")
    sheet = workbook[sheet_name]
    values = [
        cell.value
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)
        for cell in row
    ]
    non_empty_count = sum(value not in {None, ""} for value in values)
    if non_empty_count == 0:
        raise ChartContractError(f"{path} contains no source values")
    address = f"${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}"
    return SourceRange(
        input=value,
        sheet=sheet_name,
        address=address,
        formula=f"{quote_sheet_name(sheet_name)}!{address}",
        min_col=min_col,
        min_row=min_row,
        max_col=max_col,
        max_row=max_row,
        point_count=point_count,
        non_empty_count=non_empty_count,
    )


def validate_bubble_sizes(workbook: Any, source: SourceRange, path: str) -> None:
    sheet = workbook[source.sheet]
    values = [
        cell.value
        for row in sheet.iter_rows(
            min_row=source.min_row,
            max_row=source.max_row,
            min_col=source.min_col,
            max_col=source.max_col,
        )
        for cell in row
        if cell.value not in {None, ""}
    ]
    has_formula = any(isinstance(value, str) and value.startswith("=") for value in values)
    has_positive_number = any(isinstance(value, (int, float)) and not isinstance(value, bool) and value > 0 for value in values)
    if not has_formula and not has_positive_number:
        raise ChartContractError(f"{path} must contain a positive numeric or formula-driven bubble size")


def require_same_point_count(expected: SourceRange | None, actual: SourceRange, path: str) -> None:
    if expected is None or expected.point_count != actual.point_count:
        count = expected.point_count if expected else 0
        raise ChartContractError(f"{path} has {actual.point_count} points; expected {count}")


def reject_unknown_fields(value: dict[str, Any], allowed: set[str], path: str) -> None:
    unknown = sorted(set(value) - allowed)
    if unknown:
        raise ChartContractError(f"{path} contains unsupported field(s): {', '.join(unknown)}")


def require_text(value: Any, path: str, maximum: int) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ChartContractError(f"{path} must be a non-empty string")
    normalized = value.strip()
    if len(normalized) > maximum or any(ord(char) < 32 and char not in {"\t", "\n"} for char in normalized):
        raise ChartContractError(f"{path} must be at most {maximum} characters and contain no control characters")
    return normalized


def require_choice(value: Any, choices: set[str], path: str) -> str:
    if not isinstance(value, str) or value not in choices:
        raise ChartContractError(f"{path} must be one of: {', '.join(sorted(choices))}")
    return value


def require_anchor(value: Any, path: str) -> str:
    if not isinstance(value, str):
        raise ChartContractError(f"{path} must be one A1 cell")
    match = CELL_PATTERN.fullmatch(value.strip())
    if not match:
        raise ChartContractError(f"{path} must be one A1 cell")
    column = range_boundaries(f"{match.group(1)}{match.group(2)}")[0]
    row = int(match.group(2))
    if column > MAX_COLUMN or row > MAX_ROW:
        raise ChartContractError(f"{path} is outside Excel worksheet bounds")
    return f"{get_column_letter(column)}{row}"


def require_dimension(value: Any, fallback: float, minimum: float, maximum: float, path: str) -> float:
    if value is None:
        return fallback
    if isinstance(value, bool) or not isinstance(value, (int, float)) or not math.isfinite(value):
        raise ChartContractError(f"{path} must be a finite number")
    numeric = float(value)
    if numeric < minimum or numeric > maximum:
        raise ChartContractError(f"{path} must be between {minimum} and {maximum}")
    return numeric


def validate_color(value: Any, path: str) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str) or not re.fullmatch(r"#?[0-9A-Fa-f]{6}", value):
        raise ChartContractError(f"{path} must be exactly six hexadecimal digits")
    return value.lstrip("#").upper()


def apply_with_openpyxl(workbook: Any, output_path: Path, operations: list[ValidatedOperation]) -> None:
    try:
        for operation in operations:
            apply_chart_openpyxl(workbook, operation)
        workbook.save(output_path)
    finally:
        workbook.close()


def apply_chart_openpyxl(workbook: Any, operation: ValidatedOperation) -> None:
    if operation.effective_chart_type == "combo":
        chart = build_combo_chart(workbook, operation)
    else:
        chart = new_chart(operation.chart_type)
        configure_chart_type(chart, operation.chart_type, operation.grouping)
        for series_spec in operation.series:
            append_series(workbook, chart, series_spec)
        if operation.chart_type not in {"scatter", "bubble"}:
            chart.set_categories(operation.category.reference(workbook))
        apply_data_labels(chart, operation.data_labels)
    configure_chart_shell(chart, operation)
    workbook[operation.sheet].add_chart(chart, operation.anchor)


def build_combo_chart(workbook: Any, operation: ValidatedOperation) -> Any:
    grouped: list[tuple[tuple[str, bool], list[ValidatedSeries]]] = []
    for series in operation.series:
        key = (series.chart_type, series.secondary_axis)
        existing = next((item for item in grouped if item[0] == key), None)
        if existing:
            existing[1].append(series)
        else:
            grouped.append((key, [series]))
    grouped.sort(key=lambda item: item[0][1])

    charts: list[tuple[Any, bool]] = []
    secondary_index = 0
    for (chart_type, secondary_axis), series_group in grouped:
        chart = new_chart(chart_type)
        configure_chart_type(chart, chart_type, operation.grouping)
        for series in series_group:
            append_series(workbook, chart, series)
        chart.set_categories(operation.category.reference(workbook))
        apply_data_labels(chart, operation.data_labels)
        if secondary_axis:
            secondary_index += 1
            chart.y_axis.axId = 200 + secondary_index * 100
            chart.y_axis.crosses = "max"
            chart.y_axis.crossAx = chart.x_axis.axId
        charts.append((chart, secondary_axis))

    base = charts[0][0]
    for chart, _ in charts[1:]:
        base += chart
    return base


def new_chart(chart_type: str) -> Any:
    if chart_type == "line":
        return LineChart()
    if chart_type in {"bar", "column"}:
        return BarChart()
    if chart_type == "pie":
        return PieChart()
    if chart_type == "doughnut":
        return DoughnutChart()
    if chart_type == "scatter":
        return ScatterChart()
    if chart_type == "area":
        return AreaChart()
    if chart_type == "bubble":
        return BubbleChart()
    raise ChartContractError(f"unsupported chart type: {chart_type}")


def configure_chart_type(chart: Any, chart_type: str, grouping: str) -> None:
    if isinstance(chart, BarChart):
        chart.type = "bar" if chart_type == "bar" else "col"
        chart.grouping = grouping
        chart.overlap = 100 if grouping in {"stacked", "percentStacked"} else 0
    elif isinstance(chart, (LineChart, AreaChart)):
        chart.grouping = "standard" if grouping == "clustered" else grouping


def configure_chart_shell(chart: Any, operation: ValidatedOperation) -> None:
    chart.title = operation.title
    chart.style = 10
    chart.width = operation.width
    chart.height = operation.height
    if operation.legend_position == "none":
        chart.legend = None
    elif chart.legend is not None:
        chart.legend.position = {"top": "t", "bottom": "b", "left": "l", "right": "r"}[operation.legend_position]


def append_series(workbook: Any, chart: Any, series_spec: ValidatedSeries) -> None:
    values = series_spec.values.reference(workbook)
    if series_spec.chart_type == "scatter":
        if series_spec.x_values is None:
            raise ChartContractError("scatter series lost its xValuesRange after validation")
        series = Series(values, series_spec.x_values.reference(workbook), title=series_spec.name)
    elif series_spec.chart_type == "bubble":
        if series_spec.x_values is None or series_spec.sizes is None:
            raise ChartContractError("bubble series lost its xValuesRange or sizeRange after validation")
        series = Series(
            values,
            series_spec.x_values.reference(workbook),
            series_spec.sizes.reference(workbook),
            title=series_spec.name,
        )
    else:
        series = Series(values, title=series_spec.name)
    apply_series_color(series, series_spec.color, series_spec.chart_type)
    chart.series.append(series)


def apply_series_color(series: Any, color: str | None, chart_type: str) -> None:
    if color is None:
        return
    try:
        if chart_type in {"line", "scatter"}:
            series.graphicalProperties.line.solidFill = color
        else:
            series.graphicalProperties.solidFill = color
    except (AttributeError, TypeError, ValueError):
        pass


def apply_data_labels(chart: Any, enabled: bool) -> None:
    if enabled:
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True


def apply_with_excel_com(
    workbook_path: Path,
    output_path: Path,
    operations: list[ValidatedOperation],
) -> None:
    import win32com.client  # type: ignore

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    workbook = None
    try:
        workbook = excel.Workbooks.Open(str(workbook_path), UpdateLinks=0, ReadOnly=False)
        for operation in operations:
            apply_chart_com(workbook, operation)
        workbook.SaveCopyAs(str(output_path))
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        excel.Quit()


def apply_chart_com(workbook: Any, operation: ValidatedOperation) -> None:
    target_sheet = workbook.Worksheets(operation.sheet)
    anchor = target_sheet.Range(operation.anchor)
    chart_object = target_sheet.ChartObjects().Add(
        anchor.Left,
        anchor.Top,
        operation.width * 28,
        operation.height * 20,
    )
    chart = chart_object.Chart
    primary_series = next(item for item in operation.series if not item.secondary_axis)
    chart.ChartType = chart_type_com(primary_series.chart_type, operation.grouping)
    chart.HasTitle = True
    chart.ChartTitle.Text = operation.title
    chart.HasLegend = operation.legend_position != "none"
    if chart.HasLegend:
        chart.Legend.Position = {"top": -4160, "bottom": -4107, "left": -4131, "right": -4152}[operation.legend_position]
    while chart.SeriesCollection().Count:
        chart.SeriesCollection(1).Delete()
    for series_spec in operation.series:
        series = chart.SeriesCollection().NewSeries()
        series.Name = series_spec.name
        series.Values = range_com(workbook, series_spec.values)
        if operation.chart_type in {"scatter", "bubble"}:
            if series_spec.x_values is None:
                raise ChartContractError("validated scatter/bubble series has no x range")
            series.XValues = range_com(workbook, series_spec.x_values)
        else:
            series.XValues = range_com(workbook, operation.category)
        if operation.chart_type == "bubble":
            if series_spec.sizes is None:
                raise ChartContractError("validated bubble series has no size range")
            series.BubbleSizes = range_com(workbook, series_spec.sizes)
        series.ChartType = chart_type_com(series_spec.chart_type, operation.grouping)
        if series_spec.secondary_axis:
            series.AxisGroup = 2
        apply_series_color_com(series, series_spec.color, series_spec.chart_type)
        if operation.data_labels:
            series.ApplyDataLabels()


def chart_type_com(chart_type: str, grouping: str) -> int:
    if chart_type == "line":
        return 64 if grouping == "percentStacked" else 63 if grouping == "stacked" else 4
    if chart_type == "bar":
        return 59 if grouping == "percentStacked" else 58 if grouping == "stacked" else 57
    if chart_type == "column":
        return 53 if grouping == "percentStacked" else 52 if grouping == "stacked" else 51
    if chart_type == "area":
        return 77 if grouping == "percentStacked" else 76 if grouping == "stacked" else 1
    return {"pie": 5, "doughnut": -4120, "scatter": -4169, "bubble": 15}[chart_type]


def range_com(workbook: Any, source: SourceRange) -> Any:
    return workbook.Worksheets(source.sheet).Range(source.address.replace("$", ""))


def apply_series_color_com(series: Any, color: str | None, chart_type: str) -> None:
    if color is None:
        return
    red, green, blue = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    excel_rgb = red + (green << 8) + (blue << 16)
    try:
        if chart_type in {"line", "scatter"}:
            series.Format.Line.ForeColor.RGB = excel_rgb
        else:
            series.Format.Fill.ForeColor.RGB = excel_rgb
    except Exception:
        pass


def inspect_package(path: Path) -> dict[str, Any]:
    with ZipFile(path, "r") as archive:
        names = set(archive.namelist())
        chart_paths = sorted(
            (name for name in names if re.fullmatch(r"xl/charts/chart[0-9]+\.xml", name)),
            key=natural_part_key,
        )
        drawing_paths = sorted(
            (name for name in names if re.fullmatch(r"xl/drawings/drawing[0-9]+\.xml", name)),
            key=natural_part_key,
        )
        chart_objects = inspect_chart_anchors(archive, names)
        object_by_part = {item["chartPart"]: item for item in chart_objects}
        chart_parts = []
        for chart_path in chart_paths:
            metadata = inspect_chart_xml(chart_path, archive.read(chart_path))
            anchor = object_by_part.get(chart_path)
            if anchor:
                metadata.update({
                    "sheet": anchor["sheet"],
                    "anchor": anchor["anchor"],
                    "drawingPart": anchor["drawingPart"],
                })
            chart_parts.append(metadata)
        return {
            "chartParts": chart_parts,
            "drawingParts": drawing_paths,
            "chartObjects": chart_objects,
        }


def inspect_chart_anchors(archive: ZipFile, names: set[str]) -> list[dict[str, Any]]:
    if "xl/workbook.xml" not in names:
        return []
    workbook_relationships = read_relationships(archive, names, "xl/workbook.xml")
    workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    chart_objects: list[dict[str, Any]] = []
    for sheet_node in (node for node in workbook_root.iter() if local_name(node.tag) == "sheet"):
        sheet_name = sheet_node.attrib.get("name", "")
        relationship_id = sheet_node.attrib.get(RELATIONSHIP_ID)
        relationship = workbook_relationships.get(relationship_id or "")
        if not relationship:
            continue
        sheet_part = relationship["target"]
        sheet_relationships = read_relationships(archive, names, sheet_part)
        for sheet_relation in sheet_relationships.values():
            if not sheet_relation["type"].endswith("/drawing"):
                continue
            drawing_part = sheet_relation["target"]
            if drawing_part not in names:
                continue
            drawing_relationships = read_relationships(archive, names, drawing_part)
            drawing_root = ElementTree.fromstring(archive.read(drawing_part))
            for anchor_node in list(drawing_root):
                chart_node = next((node for node in anchor_node.iter() if local_name(node.tag) == "chart"), None)
                if chart_node is None:
                    continue
                chart_relationship = drawing_relationships.get(chart_node.attrib.get(RELATIONSHIP_ID, ""))
                if not chart_relationship or not chart_relationship["type"].endswith("/chart"):
                    continue
                from_node = next((node for node in anchor_node if local_name(node.tag) == "from"), None)
                anchor = None
                if from_node is not None:
                    column_node = next((node for node in from_node if local_name(node.tag) == "col"), None)
                    row_node = next((node for node in from_node if local_name(node.tag) == "row"), None)
                    if column_node is not None and row_node is not None:
                        anchor = f"{get_column_letter(int(column_node.text or '0') + 1)}{int(row_node.text or '0') + 1}"
                chart_objects.append({
                    "chartPart": chart_relationship["target"],
                    "drawingPart": drawing_part,
                    "sheet": sheet_name,
                    "anchor": anchor,
                })
    return sorted(chart_objects, key=lambda item: natural_part_key(item["chartPart"]))


def inspect_chart_xml(path: str, payload: bytes) -> dict[str, Any]:
    root = ElementTree.fromstring(payload)
    chart_elements = [local_name(node.tag) for node in root.iter() if local_name(node.tag) in CHART_ELEMENT_NAMES]
    title_node = next((node for node in root.iter() if local_name(node.tag) == "title"), None)
    title = ""
    if title_node is not None:
        title_fragments = [node.text or "" for node in title_node.iter() if local_name(node.tag) == "t"]
        if not title_fragments:
            title_fragments = [node.text or "" for node in title_node.iter() if local_name(node.tag) == "v"]
        title = "".join(title_fragments)
    legend_node = next((node for node in root.iter() if local_name(node.tag) == "legend"), None)
    legend_position = None
    if legend_node is not None:
        position_node = next((node for node in legend_node.iter() if local_name(node.tag) == "legendPos"), None)
        legend_position = position_node.attrib.get("val", "r") if position_node is not None else "r"
    sources = sorted({
        canonical_formula(node.text or "")
        for node in root.iter()
        if local_name(node.tag) == "f" and (node.text or "").strip()
    })
    return {
        "path": path,
        "sha256": hashlib.sha256(payload).hexdigest(),
        "bytes": len(payload),
        "chartElements": chart_elements,
        "barDirections": [node.attrib.get("val") for node in root.iter() if local_name(node.tag) == "barDir"],
        "title": title,
        "legendPosition": legend_position,
        "seriesCount": sum(1 for node in root.iter() if local_name(node.tag) == "ser"),
        "valueAxisCount": sum(1 for node in root.iter() if local_name(node.tag) == "valAx"),
        "sources": sources,
        "sheet": None,
        "anchor": None,
        "drawingPart": None,
    }


def read_relationships(
    archive: ZipFile,
    names: set[str],
    source_part: str,
) -> dict[str, dict[str, str]]:
    directory, filename = posixpath.split(source_part)
    relationships_part = posixpath.join(directory, "_rels", f"{filename}.rels")
    if relationships_part not in names:
        return {}
    root = ElementTree.fromstring(archive.read(relationships_part))
    relationships: dict[str, dict[str, str]] = {}
    for node in root:
        relationship_id = node.attrib.get("Id")
        target = node.attrib.get("Target")
        if not relationship_id or not target or node.attrib.get("TargetMode") == "External":
            continue
        resolved_target = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join(directory, target))
        relationships[relationship_id] = {"target": resolved_target, "type": node.attrib.get("Type", "")}
    return relationships


def verify_persisted_charts(
    before: dict[str, Any],
    after: dict[str, Any],
    operations: list[ValidatedOperation],
) -> list[dict[str, Any]]:
    expected_count = len(operations)
    before_part_count = len(before["chartParts"])
    after_part_count = len(after["chartParts"])
    if after_part_count != before_part_count + expected_count:
        raise ChartContractError(
            f"post-save package verification expected {before_part_count + expected_count} chart parts, found {after_part_count}"
        )
    before_object_count = len(before["chartObjects"])
    after_object_count = len(after["chartObjects"])
    if after_object_count != before_object_count + expected_count:
        raise ChartContractError(
            f"post-save package verification expected {before_object_count + expected_count} anchored chart objects, found {after_object_count}"
        )

    receipts: list[dict[str, Any]] = []
    for operation in operations:
        candidates = [
            item for item in after["chartParts"]
            if item.get("sheet") == operation.sheet and item.get("anchor") == operation.anchor
        ]
        if len(candidates) != 1:
            raise ChartContractError(
                f"post-save package verification found {len(candidates)} charts at {operation.sheet}!{operation.anchor}"
            )
        chart_part = candidates[0]
        verify_chart_metadata(chart_part, operation)
        receipts.append({
            **operation.receipt(),
            "chartPart": chart_part["path"],
            "drawingPart": chart_part["drawingPart"],
            "verified": True,
            "persisted": {
                "chartElements": chart_part["chartElements"],
                "barDirections": chart_part["barDirections"],
                "seriesCount": chart_part["seriesCount"],
                "valueAxisCount": chart_part["valueAxisCount"],
                "sources": chart_part["sources"],
                "sha256": chart_part["sha256"],
            },
        })
    return receipts


def verify_chart_metadata(chart_part: dict[str, Any], operation: ValidatedOperation) -> None:
    if chart_part.get("title") != operation.title:
        raise ChartContractError(
            f"post-save package verification title mismatch at {operation.sheet}!{operation.anchor}: {chart_part.get('title')!r}"
        )
    expected_legend = {"top": "t", "bottom": "b", "left": "l", "right": "r", "none": None}[operation.legend_position]
    if chart_part.get("legendPosition") != expected_legend:
        raise ChartContractError(
            f"post-save package verification legend mismatch at {operation.sheet}!{operation.anchor}"
        )
    if chart_part.get("seriesCount") != len(operation.series):
        raise ChartContractError(
            f"post-save package verification series mismatch at {operation.sheet}!{operation.anchor}"
        )

    expected_types = {chart_element_name(item.chart_type) for item in operation.series}
    actual_types = set(chart_part.get("chartElements", []))
    if not expected_types.issubset(actual_types):
        raise ChartContractError(
            f"post-save package verification chart type mismatch at {operation.sheet}!{operation.anchor}"
        )
    expected_bar_directions = {
        "bar" if item.chart_type == "bar" else "col"
        for item in operation.series
        if item.chart_type in {"bar", "column"}
    }
    if not expected_bar_directions.issubset(set(chart_part.get("barDirections", []))):
        raise ChartContractError(
            f"post-save package verification bar direction mismatch at {operation.sheet}!{operation.anchor}"
        )

    expected_sources = {canonical_formula(item.values.formula) for item in operation.series}
    if operation.effective_chart_type == "combo" or operation.chart_type not in {"scatter", "bubble"}:
        expected_sources.add(canonical_formula(operation.category.formula))
    for item in operation.series:
        if item.x_values:
            expected_sources.add(canonical_formula(item.x_values.formula))
        if item.sizes:
            expected_sources.add(canonical_formula(item.sizes.formula))
    if not expected_sources.issubset(set(chart_part.get("sources", []))):
        missing = sorted(expected_sources - set(chart_part.get("sources", [])))
        raise ChartContractError(
            f"post-save package verification missing source formula(s) at {operation.sheet}!{operation.anchor}: {', '.join(missing)}"
        )
    if any(item.secondary_axis for item in operation.series) and chart_part.get("valueAxisCount", 0) < 2:
        raise ChartContractError(
            f"post-save package verification found no secondary value axis at {operation.sheet}!{operation.anchor}"
        )


def chart_element_name(chart_type: str) -> str:
    return {
        "line": "lineChart",
        "bar": "barChart",
        "column": "barChart",
        "pie": "pieChart",
        "doughnut": "doughnutChart",
        "scatter": "scatterChart",
        "area": "areaChart",
        "bubble": "bubbleChart",
    }[chart_type]


def canonical_formula(value: str) -> str:
    stripped = value.strip().lstrip("=")
    match = RANGE_PATTERN.fullmatch(stripped)
    if not match:
        return stripped.replace("$", "").casefold()
    sheet = match.group(1).replace("''", "'") if match.group(1) else (match.group(2) or "")
    address = f"{match.group(3).upper()}{match.group(4)}:{match.group(5).upper()}{match.group(6)}"
    return f"{sheet.casefold()}!{address}"


def quote_sheet_name(name: str) -> str:
    return f"'{name.replace(chr(39), chr(39) * 2)}'"


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def natural_part_key(path: str) -> tuple[str, int]:
    match = re.search(r"([0-9]+)\.xml$", path)
    return (path.rsplit("/", 1)[0], int(match.group(1)) if match else 0)


def create_temporary_workbook_path(workbook_path: Path) -> Path:
    handle, raw_path = tempfile.mkstemp(
        prefix=f".{workbook_path.stem}.charts-",
        suffix=workbook_path.suffix,
        dir=workbook_path.parent,
    )
    os.close(handle)
    path = Path(raw_path)
    path.unlink(missing_ok=True)
    return path


def success_receipt(
    workbook_path: Path,
    engine: str,
    before: dict[str, Any],
    after: dict[str, Any],
    operations: list[ValidatedOperation],
    operation_receipts: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "schema": 2,
        "contractVersion": CONTRACT_VERSION,
        "status": "applied",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "workbook": workbook_path.name,
        "engine": engine,
        "appliedChartCount": len(operations),
        "operationCount": len(operations),
        "package": {
            "chartObjectCountBefore": len(before["chartObjects"]),
            "chartObjectCountAfter": len(after["chartObjects"]),
            "chartPartCountBefore": len(before["chartParts"]),
            "chartPartCountAfter": len(after["chartParts"]),
            "drawingPartCountAfter": len(after["drawingParts"]),
        },
        "operations": operation_receipts,
    }


def rejection_receipt(
    workbook_path: Path,
    engine: str,
    before: dict[str, Any],
    operation_count: int,
    error: Exception,
) -> dict[str, Any]:
    return {
        "schema": 2,
        "contractVersion": CONTRACT_VERSION,
        "status": "rejected",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "workbook": workbook_path.name,
        "engine": engine,
        "appliedChartCount": 0,
        "operationCount": operation_count,
        "package": {
            "chartObjectCountBefore": len(before["chartObjects"]),
            "chartObjectCountAfter": len(before["chartObjects"]),
            "chartPartCountBefore": len(before["chartParts"]),
            "chartPartCountAfter": len(before["chartParts"]),
            "drawingPartCountAfter": len(before["drawingParts"]),
        },
        "operations": [],
        "error": {"type": type(error).__name__, "message": str(error)[:1_000]},
    }


def empty_package_summary() -> dict[str, Any]:
    return {"chartParts": [], "drawingParts": [], "chartObjects": []}


def write_json_atomic(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, raw_path = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    temporary = Path(raw_path)
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            json.dump(payload, stream, indent=2, ensure_ascii=True)
            stream.write("\n")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


if __name__ == "__main__":
    raise SystemExit(main())
